import tkinter as tk
from tkinter import ttk, filedialog
import csv
import openpyxl
from datetime import datetime
from tkcalendar import DateEntry

def validar_numero(valor):
    return valor.isdigit() or valor == ""

def validar_letras(valor):
    return valor.isalpha() or valor == ""

class VentanaAdultos(tk.Frame):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)

        self.bloque_superior = ttk.Frame(self)
        self.bloque_superior.grid(row=0, column=0, padx=0, pady=0, sticky="w")

        etiquetas_superior_izquierda = ["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos"]
        etiquetas_superior_derecha = ["Fecha de Nacimiento", "Teléfono", "DNI"]
        self.entradas_superior_izquierda = []
        self.entradas_superior_derecha = []

        for i, etiqueta in enumerate(etiquetas_superior_izquierda):
            etiqueta_widget = ttk.Label(self.bloque_superior, text=etiqueta, anchor="w")
            etiqueta_widget.grid(row=i, column=0, pady=5, padx=10, sticky="w")

            if etiqueta == "Sexo":
                opciones_sexo = ["Masculino", "Femenino"]
                self.sexo_var = tk.StringVar()
                combo_sexo = ttk.Combobox(self.bloque_superior, textvariable=self.sexo_var, values=opciones_sexo, state="readonly")
                combo_sexo.grid(row=i, column=1, pady=5, padx=10, sticky="w")
                self.entradas_superior_izquierda.append(combo_sexo)
            elif etiqueta in ["Edad"]:
                entrada_widget = ttk.Entry(self.bloque_superior, validate="key")
                entrada_widget.configure(validatecommand=(entrada_widget.register(validar_numero), "%P"))
                entrada_widget.grid(row=i, column=1, pady=5, padx=10, sticky="w")
                self.entradas_superior_izquierda.append(entrada_widget)
            elif etiqueta in ["Fecha", "Fecha de Nacimiento"]:
                entrada_widget = DateEntry(self.bloque_superior, date_pattern='dd-mm-yyyy', locale='es_ES')
                entrada_widget.grid(row=i, column=1, pady=5, padx=10, sticky="w")
                self.entradas_superior_izquierda.append(entrada_widget)
            elif etiqueta in ["Nombres", "Apellidos"]:
                entrada_widget = ttk.Entry(self.bloque_superior, validate="key")
                entrada_widget.configure(validatecommand=(entrada_widget.register(validar_letras), "%P"))
                entrada_widget.grid(row=i, column=1, pady=5, padx=10, sticky="w")
                self.entradas_superior_izquierda.append(entrada_widget)
            else:
                entrada_widget = ttk.Entry(self.bloque_superior)
                entrada_widget.grid(row=i, column=1, pady=5, padx=10, sticky="w")
                self.entradas_superior_izquierda.append(entrada_widget)

        for i, etiqueta in enumerate(etiquetas_superior_derecha):
            etiqueta_widget = ttk.Label(self.bloque_superior, text=etiqueta, anchor="w")
            etiqueta_widget.grid(row=i, column=2, pady=5, padx=10, sticky="w")

            if etiqueta == "Fecha de Nacimiento":
                entrada_widget = DateEntry(self.bloque_superior, date_pattern='dd-mm-yyyy', locale='es_ES')
            elif etiqueta in ["Edad", "DNI", "Teléfono"]:
                entrada_widget = ttk.Entry(self.bloque_superior, validate="key")
                entrada_widget.configure(validatecommand=(entrada_widget.register(validar_numero), "%P"))
            else:
                entrada_widget = ttk.Entry(self.bloque_superior)

            entrada_widget.grid(row=i, column=3, pady=5, padx=10, sticky="w")
            self.entradas_superior_derecha.append(entrada_widget)


        self.boton_guardar = ttk.Button(self.bloque_superior, text="Guardar", command=self.guardar_datos)
        self.boton_guardar.grid(row=len(etiquetas_superior_izquierda), columnspan=4, pady=20, padx=(10, 0), sticky="w")

        self.boton_exportar_excel = ttk.Button(self.bloque_superior, text="Exportar a Excel", command=self.exportar_a_excel)
        self.boton_exportar_excel.grid(row=len(etiquetas_superior_izquierda), column=4, pady=20, padx=(10, 0), sticky="w")

        self.boton_importar_excel = ttk.Button(self.bloque_superior, text="Importar desde Excel", command=self.importar_excel)
        self.boton_importar_excel.grid(row=len(etiquetas_superior_izquierda), column=5, pady=20, padx=(10, 0), sticky="w")

        self.boton_modificar = ttk.Button(self.bloque_superior, text="Modificar Seleccionado", command=self.modificar_seleccion)
        self.boton_modificar.grid(row=len(etiquetas_superior_izquierda), column=3, pady=20, padx=(0, 10), sticky="e")

        self.boton_eliminar = ttk.Button(self.bloque_superior, text="Eliminar Seleccionado", command=self.eliminar_seleccion)
        self.boton_eliminar.grid(row=len(etiquetas_superior_izquierda), column=2, pady=20, padx=(0, 10), sticky="e")

        self.etiqueta_buscar_hc = ttk.Label(self.bloque_superior, text="Buscar por HC:")
        self.etiqueta_buscar_hc.grid(row=len(etiquetas_superior_izquierda)+1, column=0, pady=5, padx=10, sticky="e")

        self.entrada_buscar_hc = ttk.Entry(self.bloque_superior)
        self.entrada_buscar_hc.grid(row=len(etiquetas_superior_izquierda)+1, column=1, pady=5, padx=10, sticky="w")

        self.etiqueta_buscar_dni = ttk.Label(self.bloque_superior, text="Buscar por DNI:")
        self.etiqueta_buscar_dni.grid(row=len(etiquetas_superior_izquierda)+2, column=0, pady=5, padx=10, sticky="e")

        self.entrada_buscar_dni = ttk.Entry(self.bloque_superior)
        self.entrada_buscar_dni.grid(row=len(etiquetas_superior_izquierda)+2, column=1, pady=5, padx=10, sticky="w")

        self.boton_buscar_hc = ttk.Button(self.bloque_superior, text="Buscar", command=self.filtrar_tabla)
        self.boton_buscar_hc.grid(row=len(etiquetas_superior_izquierda)+1, column=2, pady=5, padx=10, sticky="w")

        self.boton_buscar_dni = ttk.Button(self.bloque_superior, text="Buscar", command=self.filtrar_tabla)
        self.boton_buscar_dni.grid(row=len(etiquetas_superior_izquierda)+2, column=2, pady=5, padx=10, sticky="w")

        self.boton_limpiar_busqueda = ttk.Button(self.bloque_superior, text="Limpiar Búsqueda", command=self.limpiar_campos_busqueda)
        self.boton_limpiar_busqueda.grid(row=len(etiquetas_superior_izquierda)+1, column=3, pady=5, padx=10, sticky="e")

        self.datos_originales = []

        self.bloque_inferior = ttk.Frame(self, borderwidth=2, relief="groove", padding=(5, 5, 5, 5))
        self.bloque_inferior.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        etiquetas_tabla = ["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"]
        self.tabla = ttk.Treeview(self.bloque_inferior, columns=etiquetas_tabla, show="headings", height=5)

        for etiqueta in etiquetas_tabla:
            self.tabla.heading(etiqueta, text=etiqueta)
            self.tabla.column(etiqueta, width=80)

        self.tabla.grid(row=0, column=0, columnspan=4, pady=10, padx=10, sticky="w")

        scrollbar_y = ttk.Scrollbar(self.bloque_inferior, orient="vertical", command=self.tabla.yview)
        scrollbar_y.grid(row=0, column=4, pady=10, padx=(0, 10), sticky="ns")
        self.tabla.configure(yscrollcommand=scrollbar_y.set)

        self.cargar_datos()

    def guardar_datos(self):
        datos = [entrada.get() for entrada in self.entradas_superior_izquierda + self.entradas_superior_derecha]
        self.tabla.insert("", "end", values=datos, tags=(datos[0],))
        self.guardar_en_csv(datos)

        for entrada in self.entradas_superior_izquierda + self.entradas_superior_derecha:
            entrada.delete(0, tk.END)

    def guardar_en_csv(self, datos):
        with open("datos_adultos.csv", mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(datos)

    def cargar_datos(self):
        try:
            with open("datos_adultos.csv", mode="r", encoding="utf-8") as file:
                reader = csv.reader(file)
                next(reader, None)
                for row in reader:
                    self.tabla.insert("", "end", values=row, tags=(row[0],))
                    self.datos_originales.append(row)
        except FileNotFoundError:
            pass

    def eliminar_seleccion(self):
        item_seleccionado = self.tabla.selection()
        if item_seleccionado:
            hc_seleccionado = self.tabla.item(item_seleccionado)['values'][0]
            self.tabla.tag_configure(hc_seleccionado, background="")
            self.tabla.delete(item_seleccionado)
            self.actualizar_archivo_csv()

    def modificar_seleccion(self):
        item_seleccionado = self.tabla.selection()
        if item_seleccionado:
            valores_seleccionados = self.tabla.item(item_seleccionado)['values']

            for i, valor in enumerate(valores_seleccionados):
                if i < len(self.entradas_superior_izquierda):
                    self.entradas_superior_izquierda[i].delete(0, tk.END)
                    self.entradas_superior_izquierda[i].insert(0, valor)
                else:
                    self.entradas_superior_derecha[i - len(self.entradas_superior_izquierda)].delete(0, tk.END)
                    self.entradas_superior_derecha[i - len(self.entradas_superior_izquierda)].insert(0, valor)

            self.tabla.delete(item_seleccionado)
            self.actualizar_archivo_csv()

    def filtrar_tabla(self):
        hc_buscar = self.entrada_buscar_hc.get()
        dni_buscar = self.entrada_buscar_dni.get()

        self.tabla.delete(*self.tabla.get_children())

        for datos_fila in self.datos_originales:
            hc_fila = datos_fila[0]
            dni_fila = datos_fila[8]

            if ((not hc_buscar or hc_fila == hc_buscar) and 
                (not dni_buscar or dni_fila == dni_buscar)):
                self.tabla.insert("", "end", values=datos_fila, tags=(datos_fila[0],))

    def limpiar_campos_busqueda(self):
        self.entrada_buscar_hc.delete(0, tk.END)
        self.entrada_buscar_dni.delete(0, tk.END)
        self.filtrar_tabla()

    def restablecer_tabla(self):
        self.tabla.selection_remove(self.tabla.selection())
        for i, fila_tabla in enumerate(self.tabla.get_children()):
            self.tabla.reattach(fila_tabla, "", i)

    def actualizar_archivo_csv(self):
        with open("datos_adultos.csv", mode="w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"])
            for fila_tabla in self.tabla.get_children():
                datos_fila = self.tabla.item(fila_tabla)['values']
                writer.writerow(datos_fila)

    def exportar_a_excel(self):
        archivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])

        if archivo_excel:
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo_excel = f"{fecha_actual}_{archivo_excel.replace('/', '_').replace(':', '')}"  # Corregir la formación de la ruta

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            encabezados = ["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"]
            sheet.append(encabezados)

            for fila_tabla in self.tabla.get_children():
                datos_fila = self.tabla.item(fila_tabla)['values']
                sheet.append(datos_fila)

            workbook.save(archivo_excel)

    def importar_excel(self):
        archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx","*csv")])

        if archivo_excel:
            workbook = openpyxl.load_workbook(archivo_excel)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tabla.insert("", "end", values=row, tags=(row[0],))
                self.guardar_en_csv(row)

            workbook.close()

            # Seleccionar la última fila insertada después de importar
            ultima_fila = self.tabla.get_children()[-1]
            self.tabla.selection_set(ultima_fila)
            
            # Llamar a la función de selección de fila después de importar
            self.seleccionar_fila_desde_importar()

    def seleccionar_fila_desde_importar(self):
        item_seleccionado = self.tabla.selection()
        if item_seleccionado:
            valores_seleccionados = self.tabla.item(item_seleccionado)['values']

            for i, valor in enumerate(valores_seleccionados):
                if i < len(self.entradas_superior_izquierda):
                    self.entradas_superior_izquierda[i].delete(0, tk.END)
                    self.entradas_superior_izquierda[i].insert(0, valor)
                else:
                    self.entradas_superior_derecha[i - len(self.entradas_superior_izquierda)].delete(0, tk.END)
                    self.entradas_superior_derecha[i - len(self.entradas_superior_izquierda)].insert(0, valor)

if __name__ == "__main__":
    ventana_adultos = VentanaAdultos()
    
    ventana_adultos.mainloop()